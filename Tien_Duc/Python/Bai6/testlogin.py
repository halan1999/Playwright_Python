import openpyxl
workbook = openpyxl.load_workbook("/Users/ducnt/Documents/Python/test excel python.xlsx")
sheet = workbook.active
#valuec2 = sheet["C2"].value
#print(valuec2)
for row in sheet.iter_rows(min_row=1, values_only=True):
    stt, username, password, result = row
    print(row)
#columm_a = [sheet.cell(row= r, column = 4).value for r in range(1, sheet.max_row + 1)]